import os, io, requests
from flask import Flask, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))

NOTION_API_KEY = os.environ['NOTION_API_KEY']
DATABASE_ID = '35eafa74cfc980d092d0e80644bd6be7'
SKIP_COLUMNS = {'Relancer'}

app = Flask(__name__)

def fetch_all_pages():
    url = f'https://api.notion.com/v1/databases/{DATABASE_ID}/query'
    headers = {
        'Authorization': f'Bearer {NOTION_API_KEY}',
        'Notion-Version': '2022-06-28',
        'Content-Type': 'application/json',
    }
    results, cursor = [], None
    while True:
        body = {'page_size': 100}
        if cursor:
            body['start_cursor'] = cursor
        r = requests.post(url, headers=headers, json=body)
        r.raise_for_status()
        data = r.json()
        results.extend(data.get('results', []))
        if not data.get('has_more'):
            break
        cursor = data.get('next_cursor')
    return results

def prop_to_str(prop):
    t = prop.get('type')
    if t == 'title':
        return ''.join(r['plain_text'] for r in prop.get('title', []))
    if t == 'rich_text':
        return ''.join(r['plain_text'] for r in prop.get('rich_text', []))
    if t == 'number':
        v = prop.get('number')
        return str(v) if v is not None else ''
    if t == 'select':
        s = prop.get('select')
        return s['name'] if s else ''
    if t == 'multi_select':
        return ', '.join(s['name'] for s in prop.get('multi_select', []))
    if t == 'status':
        s = prop.get('status')
        return s['name'] if s else ''
    if t == 'checkbox':
        return 'Oui' if prop.get('checkbox') else 'Non'
    if t == 'date':
        d = prop.get('date')
        if not d:
            return ''
        start = d.get('start', '')
        end = d.get('end', '')
        return f"{start} ? {end}" if end else start
    if t == 'email':
        return prop.get('email') or ''
    if t == 'phone_number':
        return prop.get('phone_number') or ''
    if t == 'url':
        return prop.get('url') or ''
    if t == 'people':
        return ', '.join(p.get('name', '') for p in prop.get('people', []))
    if t == 'files':
        return ', '.join(
            f.get('name', '') for f in prop.get('files', [])
        )
    if t == 'formula':
        f = prop.get('formula', {})
        ft = f.get('type')
        if ft == 'string':
            return f.get('string') or ''
        if ft == 'number':
            v = f.get('number')
            return str(v) if v is not None else ''
        if ft == 'boolean':
            return 'Oui' if f.get('boolean') else 'Non'
        if ft == 'date':
            d = f.get('date')
            return d.get('start', '') if d else ''
    if t == 'rollup':
        ro = prop.get('rollup', {})
        if ro.get('type') == 'array':
            parts = [prop_to_str(item) for item in ro.get('array', [])]
            return ', '.join(p for p in parts if p)
        if ro.get('type') == 'number':
            v = ro.get('number')
            return str(v) if v is not None else ''
    if t == 'relation':
        return str(len(prop.get('relation', []))) + ' li?(s)'
    if t == 'created_time':
        return (prop.get('created_time') or '')[:10]
    if t == 'last_edited_time':
        return (prop.get('last_edited_time') or '')[:10]
    if t == 'created_by':
        return prop.get('created_by', {}).get('name', '')
    if t == 'last_edited_by':
        return prop.get('last_edited_by', {}).get('name', '')
    return ''

@app.route('/export-eleves')
def export_eleves():
    pages = fetch_all_pages()

    # Collect ordered column names (title col first, then rest)
    all_keys = {}
    for page in pages:
        for k in page.get('properties', {}):
            if k not in SKIP_COLUMNS:
                all_keys[k] = True
    columns = list(all_keys.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = 'Eleves'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    for row_idx, page in enumerate(pages, 2):
        props = page.get('properties', {})
        for col_idx, col_name in enumerate(columns, 1):
            prop = props.get(col_name, {})
            value = prop_to_str(prop) if prop else ''
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        as_attachment=True,
        download_name='eleves.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

if __name__ == '__main__':
    app.run(host='127.0.0.1', port=5003)
