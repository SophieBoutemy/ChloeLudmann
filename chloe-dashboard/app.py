import io
import json
import os
import re
import smtplib
import sqlite3
import hashlib
import hmac as _hmac
import subprocess
import uuid
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps

import anthropic

import requests
from dotenv import load_dotenv, set_key
from flask import Flask, jsonify, render_template, redirect, request, send_file, session, url_for
from itsdangerous import BadSignature, SignatureExpired, URLSafeTimedSerializer
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import csv
from werkzeug.middleware.proxy_fix import ProxyFix

DOTENV_PATH = '/home/ubuntu/automations/.env'
AUTOMATIONS_CONFIG = os.path.join(os.path.dirname(__file__), 'automations.json')

load_dotenv(DOTENV_PATH)

app = Flask(__name__)
app.secret_key = os.getenv('DASHBOARD_SECRET_KEY', 'changeme-please-set-in-env')
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_prefix=1)

DASHBOARD_USER     = os.getenv('DASHBOARD_USER', 'admin')
DASHBOARD_PASSWORD = os.getenv('DASHBOARD_PASSWORD', 'admin')

SMTP_HOST = 'ssl0.ovh.net'
SMTP_PORT = 465

NOTION_API_KEY      = os.getenv('NOTION_API_KEY', '')
DRIVE_FACTURES_URL  = os.getenv('DRIVE_FACTURES_URL', '')
NOTION_ELEVES_DB   = '35eafa74cfc980d092d0e80644bd6be7'
NOTION_FACTURES_DB = '327afa74cfc980328301eec9bb7996e5'
NOTION_BASE        = 'https://api.notion.com/v1'
FACTURES_SCRIPT    = '/home/ubuntu/automations/factures/factures.py'

DOCAGE_EMAIL   = os.getenv('DOCAGE_EMAIL', '')
DOCAGE_API_KEY = os.getenv('DOCAGE_API_KEY', '')
DOCAGE_BASE    = 'https://api.docage.com'

WAITLIST_JSON    = '/home/ubuntu/automations/liste_attente/waitlist.json'
WAITLIST_LOG     = '/home/ubuntu/automations/logs/liste_attente.log'
WAITLIST_SERVICE = 'liste-attente.service'

RECURRENCE_DB_PATH     = '/home/ubuntu/automations/recurrence_calendly/pending.db'
TALLY_FORM_URL         = 'https://tally.so/r/VLqbG6'
LOG_RECURRENCE_WEBHOOK = '/home/ubuntu/automations/logs/recurrence_calendly.log'
LOG_RECURRENCE_RETRY   = '/home/ubuntu/automations/logs/recurrence_retry.log'

RGPD_PURGE_LOG = '/home/ubuntu/automations/logs/purge_rgpd.log'
RGPD_LOG_FILES = [
    LOG_RECURRENCE_WEBHOOK,
    LOG_RECURRENCE_RETRY,
]

ANTHROPIC_API_KEY = os.getenv('ANTHROPIC_API_KEY', '')
BTP_DB_PATH       = os.path.join(os.path.dirname(__file__), 'brief_to_post.db')

# ── Adjoint Prospection ───────────────────────────────────────────────────────
_DASH_DIR    = os.path.dirname(__file__)
SECTEURS_PATH = os.path.join(_DASH_DIR, 'secteurs_naf.json')
QUOTA_PATH    = os.path.join(_DASH_DIR, 'prospection', 'email_quota.json')

with open(SECTEURS_PATH, encoding='utf-8') as _sf:
    SECTEURS_NAF = json.load(_sf)

STATUT_OPTIONS = [
    ('désinscrit',          'Désinscrit'),
    ('proposition envoyée', 'Proposition envoyée'),
    ('rendez-vous fixé',    'Rendez-vous fixé'),
    ('injoignable',         'Injoignable'),
    ('pas intéressé',       'Pas intéressé'),
    ('archivé',             'Archivé'),
]

from prospection.storage import (
    EXPORT_COLS, delete_contacts, delete_contacts_empty,
    get_all_for_export, get_contact_by_id, get_contacts,
    get_filtered_for_export, get_funnel_stats, get_naf_list,
    get_priorite_contacts, get_stats, get_today_contacts,
    get_tranche_list, init_db as init_prospection_db,
    is_desinscrit, update_contact, upsert_contact,
)
from prospection.runner import get_task_status, start_search, start_search_ch, start_search_be
from prospection.drafter import EMAIL_TYPES, generate_draft, load_profile, save_profile
from prospection.scraper import TRANCHE_LABELS

init_prospection_db()


def _get_quota():
    today = datetime.now().strftime('%Y-%m-%d')
    try:
        with open(QUOTA_PATH) as f:
            q = json.load(f)
        if q.get('date') == today:
            return q.get('count', 0)
    except Exception:
        pass
    return 0


def _increment_quota():
    today = datetime.now().strftime('%Y-%m-%d')
    count = _get_quota() + 1
    with open(QUOTA_PATH, 'w') as f:
        json.dump({'date': today, 'count': count}, f)
    return count


BTP_SYSTEM_PROMPT = """Tu es expert en communication réseaux sociaux.

Voici trois exemples de publications de cette auteure. Respecte scrupuleusement son style, son rythme, son ton, ses émojis et sa façon d'écrire :

Exemple 1 : J'attendais juste d'être clouée au lit pour vous dévoiler les 548 trésors que j'ai filmés depuis septembre 2025 🤧😁 En parallèle des cours de chant, il se passe tellement de choses ! Beaucoup d'élèves ont osé monter sur scène. Et c'était incroyable 🥰 ILS L'ONT FAIT 🔥 #chantrennes #osersavoix #chantersurscène #agilitévocale #vaincresapeur

Exemple 2 : ✨ Il y a 1 an, 20 voix se sont unies… et aujourd'hui, notre chorale souffle sa première bougie avec toujours autant d'énergie et de passion ! 🎶 A cappella, body-percussions contagieuses et répertoire folk 🎤🔥 Nous accueillons avec joie de nouveaux choristes prêts à partager l'aventure. Et nous rêvons déjà de faire résonner nos harmonies en CONCERT à Rennes 🎵💫 Rejoins-nous ❤️ #chantrennes #choralerennes #folk #folkseventies #bodypercussions

Exemple 3 : Nouvelle rentrée, nouveau challenge : m'exposer un peu plus sur les réseaux et parler de sujets sensibles sur la voix et la technique vocale. 💪🤓 Ici, pas de montage, pas de mixage, pas de micro, pas d'effets. Plein d'imperfections, de la fatigue aussi un peu, de la peur d'être jugée. Jeudi prochain, je passe devant le jury pour candidater au diplôme d'Etat de prof de chant et de technique vocale. Je dois chanter une chanson et j'ai choisi Wicked Game de Chris Isaak. Pour cell.eux qui ne le savent pas encore, je sors d'une longue période de dysphonie de tensions avec des nodules "kisses" (le seul aspect mignon de ce diag'). Plus d'aigus, des douleurs dans la gorge, des yodels incontrôlés... J'ai d'abord appris à trouver des compensations me permettant de continuer à chanter et à enseigner : baisser la tona', ne chanter que dans les graves. Mais à la longue, elles n'ont fait qu'alimenter le problème. Je me suis renseignée, j'ai potassé, je me suis entourée de spécialistes qui m'ont beaucoup aidée (vous-mêmes vous savez). J'ai aussi été soutenue tellement fort par mes élèves 💛 Je me suis formée avec @chant_voix_corps (meilleure décision ever), pour comprendre combien de nombreux mythes et fausses croyances au sujet de la technique vocale sont responsables de ma situation - et de celles d'élèves que j'accompagne également. Aujourd'hui, je suis fière de pouvoir à nouveau m'amuser avec les aigus, décider des yodels, et chanter fort (ou plutôt, avec résonance 🤓). Mais le stress emmène parfois avec lui ces anciennes petites tensions musculaires, qui ne m'aident pas. C'est aussi le traumatisme d'une voix qui déraille de façon incontrôlée, alors que je suis prof' de chant (#crisedelégitimité) qui s'exprime par anticipation. Alors chanter devant vous ici sans filtre, c'est un peu comme me préparer à chanter devant le jury. 🥹 70% des professionnel.les de la voix ont ou ont connu une pathologie vocale (même bénigne). Il faut en parler ! Et montrer que c'est possible de poursuivre sa carrière, de s'ajuster, de comprendre, de s'en sortir et de vibrer encore plus qu'avant. #oserchanter #dysphonie #techniquevocale

RÈGLES DE TON
- Professionnel : ton factuel, vocabulaire métier, phrases complètes, aucune familiarité excessive
- Émouvant : expérience personnelle, vocabulaire sensoriel, phrases parfois courtes, une émotion dominante
- Informatif : structure claire, pédagogie, faits, neutralité émotionnelle
- Inspirant : projection positive, énergie, encouragement, appel à l'action mesuré

RÈGLES PAR RÉSEAU
Règle générale : toujours privilégier la lisibilité sur mobile. Utiliser des paragraphes courts et espacés.

Instagram :
texte court à moyen, paragraphes très courts (1 à 2 phrases max), rythme visuel aéré,
ton naturel et incarné, emojis autorisés si cohérents, première phrase accrocheuse,
appel à l'action simple possible, exactement 5 hashtags en fin de publication

Facebook :
texte narratif plus développé, paragraphes de 2 à 4 phrases, style storytelling,
ton chaleureux et humain, emojis sobres, question ou invitation à réagir possible en fin,
hashtags facultatifs (0 à 2 max) en fin de publication

LinkedIn :
paragraphes courts (1 à 3 phrases max), première ligne accrocheuse, ton professionnel
et crédible mais humain, partage d'expérience/réflexion, 0 à 2 emojis max, pas de hashtags
au milieu, maximum 3 hashtags en fin, terminer par une ouverture/question si possible

Newsletter :
structure email claire, introduction courte et accrocheuse (1 paragraphe), développement
en 4 à 6 paragraphes substantiels (3 à 5 phrases chacun), pas de hashtags, appel à l'action
final possible, pas d'excès d'emojis, longueur cible 400 à 600 mots. Rendre le résultat en
UN SEUL bloc de texte continu (objet + contenu ensemble), sans séparation ni parsing distinct.

MOTS-CLÉS À PRIVILÉGIER
- développer sa musicalité
Utilise ce mot-clé de façon naturelle dans le texte, intégré au discours. Ne le répète pas
inutilement. Ne le liste pas. Ne le transforme pas en hashtag sauf demande explicite.

RÈGLES CRITIQUES
Le contenu doit être basé uniquement sur le brief utilisateur. Les exemples servent
uniquement au style, jamais au contenu. Si un élément n'est pas dans le brief, ne pas l'inventer.

Si le brief est vide, absent, tronqué ou incohérent, retourne uniquement :
"Erreur : brief invalide ou mal formaté."

---
Maintenant génère un contenu pour {reseau} avec ce brief :

{brief}

Ton souhaité : {ton}.

Retourne uniquement le texte final complet, prêt à être publié ou copié-collé, sans
préambule, sans JSON, sans balises."""

# Maps frontend field key → (Notion property name, Notion type)
FIELD_MAP = {
    'nom':            ('Nom complet',           'rich_text'),
    'type_eleve':     ("Type d'élève",          'multi_select'),
    'rattrapage':     ('Rattrapage',            'rich_text'),
    'infos':          ('Infos',                 'rich_text'),
    'infos_calendly': ('Infos Calendly',        'rich_text'),
    'boite_mail':     ('Boîte mail',            'select'),
    'statut_contrat': ('Statut contrat envoyé', 'select'),
}


# ── Auth helpers ──────────────────────────────────────────────────────────────

def get_serializer():
    return URLSafeTimedSerializer(app.secret_key)


def _make_unsub_token(contact_id):
    sig = _hmac.new(app.secret_key.encode(), str(contact_id).encode(), hashlib.sha256).hexdigest()[:8]
    return f"{contact_id}-{sig}"


def _verify_unsub_token(token):
    try:
        contact_id_str, sig = token.rsplit('-', 1)
        contact_id = int(contact_id_str)
        expected = _hmac.new(app.secret_key.encode(), str(contact_id).encode(), hashlib.sha256).hexdigest()[:8]
        if _hmac.compare_digest(sig, expected):
            return contact_id
    except Exception:
        pass
    return None


def _plain_to_html(text):
    """Convert plain-text email body to minimal HTML with a clickable unsubscribe link."""
    import re as _re
    safe = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
    safe = _re.sub(
        r'Pour ne plus recevoir de messages\s*:\s*(https?://\S+)',
        '<a href="\\1" style="color:#666666;text-decoration:none">Se désinscrire</a>',
        safe,
    )
    safe = _re.sub(
        r'(https?://mon-adjoint-ia\.fr/rgpd\.html)',
        '<a href="\\1" style="color:#666666;text-decoration:none">\\1</a>',
        safe,
    )
    safe = _re.sub(
        r'(https?://)?mon-adjoint-ia\.fr(?![/\w])',
        '<a href="https://mon-adjoint-ia.fr" style="color:#354626;text-decoration:underline">mon-adjoint-ia.fr</a>',
        safe,
    )
    safe = safe.replace('\n', '<br>\n')
    return (
        '<!DOCTYPE html><html><head><meta charset="UTF-8"></head>'
        '<body style="font-family:sans-serif;max-width:600px;margin:0 auto;'
        'padding:24px;line-height:1.7;color:#222;font-size:15px">'
        + safe + '</body></html>'
    )


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            if request.is_json:
                return jsonify({'error': 'Session expirée, veuillez vous reconnecter.'}), 401
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ── Dashboard helpers ─────────────────────────────────────────────────────────

def _last_execution(log_path, stale_after_hours=24):
    try:
        mtime = os.path.getmtime(log_path)
        last = datetime.fromtimestamp(mtime)
        delta = datetime.now() - last
        minutes = int(delta.total_seconds() / 60)
        if minutes < 60:
            label = f'il y a {minutes} min'
        elif minutes < 1440:
            label = f'il y a {minutes // 60}h'
        else:
            label = f'il y a {minutes // 1440}j'
        stale_seconds = stale_after_hours * 3600
        if delta.total_seconds() < stale_seconds * 0.5:
            status = 'ok'
        elif delta.total_seconds() < stale_seconds:
            status = 'warn'
        else:
            status = 'stale'
        return {'label': label, 'status': status}
    except FileNotFoundError:
        return {'label': 'Aucune exécution enregistrée', 'status': 'unknown'}
    except Exception:
        return {'label': 'Erreur de lecture', 'status': 'unknown'}


def _systemd_status(service):
    try:
        result = subprocess.run(
            ['systemctl', 'is-active', service],
            capture_output=True, text=True, timeout=5
        )
        state = result.stdout.strip()
        if state == 'active':
            return {'label': 'En ligne', 'status': 'ok'}
        elif state == 'inactive':
            return {'label': 'Inactif', 'status': 'warn'}
        elif state == 'failed':
            return {'label': 'Erreur', 'status': 'stale'}
        else:
            return {'label': state, 'status': 'unknown'}
    except Exception:
        return {'label': 'Inconnu', 'status': 'unknown'}


# ── Notion helpers ────────────────────────────────────────────────────────────

def _notion_headers():
    return {
        'Authorization': f'Bearer {NOTION_API_KEY}',
        'Notion-Version': '2022-06-28',
        'Content-Type': 'application/json',
    }


def _notion_prop(prop):
    t = prop.get('type', '')
    if t == 'title':
        return ''.join(r['plain_text'] for r in prop.get('title', []))
    if t == 'rich_text':
        return ''.join(r['plain_text'] for r in prop.get('rich_text', []))
    if t == 'select':
        s = prop.get('select')
        return s['name'] if s else ''
    if t == 'multi_select':
        return ', '.join(s['name'] for s in prop.get('multi_select', []))
    if t == 'date':
        d = prop.get('date')
        return d['start'] if d else ''
    if t == 'email':
        return prop.get('email') or ''
    if t == 'checkbox':
        return 'Oui' if prop.get('checkbox') else ''
    if t == 'number':
        v = prop.get('number')
        return str(v) if v is not None else ''
    if t == 'url':
        return prop.get('url') or ''
    if t == 'phone_number':
        return prop.get('phone_number') or ''
    return ''


def _build_notion_prop(field_type, value):
    if field_type == 'rich_text':
        return {'rich_text': [{'text': {'content': value}}] if value else []}
    if field_type == 'select':
        return {'select': {'name': value} if value else None}
    if field_type == 'multi_select':
        names = [v.strip() for v in value.split(',') if v.strip()]
        return {'multi_select': [{'name': n} for n in names]}
    return None


def _page_to_eleve(page):
    p = page.get('properties', {})
    return {
        'id':              page['id'],
        'notion_url':      page.get('url', ''),
        'email':           _notion_prop(p.get('Email', {})),
        'nom':             _notion_prop(p.get('Nom complet', {})),
        'type_eleve':      _notion_prop(p.get("Type d'élève", {})),
        'rattrapage':      _notion_prop(p.get('Rattrapage', {})),
        'infos':           _notion_prop(p.get('Infos', {})),
        'boite_mail':      _notion_prop(p.get('Boîte mail', {})),
        'date_mail':       _notion_prop(p.get('Date du mail', {})),
        'infos_calendly':  _notion_prop(p.get('Infos Calendly', {})),
        'resume_mail':     _notion_prop(p.get('Résumé du mail', {})),
        'date_newsletter': _notion_prop(p.get('Date Newsletter envoyée', {})),
        'date_contrat':    _notion_prop(p.get('Date contrat envoyé', {})),
        'statut_contrat':  _notion_prop(p.get('Statut contrat envoyé', {})),
        'date_relance':    _notion_prop(p.get('Date de relance', {})),
    }


def fetch_eleves():
    url = f'{NOTION_BASE}/databases/{NOTION_ELEVES_DB}/query'
    headers = _notion_headers()
    results, cursor = [], None
    while True:
        body = {
            'page_size': 100,
            'sorts': [{'property': 'Nom complet', 'direction': 'ascending'}],
        }
        if cursor:
            body['start_cursor'] = cursor
        r = requests.post(url, headers=headers, json=body, timeout=15)
        r.raise_for_status()
        data = r.json()
        results.extend(data.get('results', []))
        if not data.get('has_more'):
            break
        cursor = data.get('next_cursor')
    return [_page_to_eleve(p) for p in results]


def fetch_eleve(page_id):
    hdrs = _notion_headers()
    r = requests.get(f'{NOTION_BASE}/pages/{page_id}', headers=hdrs, timeout=10)
    r.raise_for_status()
    eleve = _page_to_eleve(r.json())

    # Fetch page body → notes
    try:
        br = requests.get(f'{NOTION_BASE}/blocks/{page_id}/children', headers=hdrs, timeout=10)
        br.raise_for_status()
        lines = []
        for block in br.json().get('results', []):
            btype = block.get('type', '')
            rich = block.get(btype, {}).get('rich_text', [])
            lines.append(''.join(rt['plain_text'] for rt in rich))
        eleve['notes'] = '\n'.join(lines)
    except Exception:
        eleve['notes'] = ''

    return eleve


def fetch_db_select_options():
    """Return select/multi_select options from the Notion database schema."""
    r = requests.get(f'{NOTION_BASE}/databases/{NOTION_ELEVES_DB}', headers=_notion_headers(), timeout=10)
    r.raise_for_status()
    props = r.json().get('properties', {})
    return {
        'boite_mail':     [o['name'] for o in props.get('Boîte mail', {}).get('select', {}).get('options', [])],
        'statut_contrat': [o['name'] for o in props.get('Statut contrat envoyé', {}).get('select', {}).get('options', [])],
        'type_eleve':     [o['name'] for o in props.get("Type d'élève", {}).get('multi_select', {}).get('options', [])],
    }


def update_notion_statut(page_id, statut):
    body = {
        'properties': {
            'Statut contrat envoyé': {'select': {'name': statut}},
            'Date de relance':       {'date': {'start': datetime.now().date().isoformat()}},
        }
    }
    r = requests.patch(f'{NOTION_BASE}/pages/{page_id}', headers=_notion_headers(), json=body, timeout=10)
    r.raise_for_status()


def save_page_notes(page_id, content):
    hdrs = _notion_headers()
    # Delete existing blocks
    existing = requests.get(f'{NOTION_BASE}/blocks/{page_id}/children', headers=hdrs, timeout=10)
    existing.raise_for_status()
    for block in existing.json().get('results', []):
        try:
            requests.delete(f'{NOTION_BASE}/blocks/{block["id"]}', headers=hdrs, timeout=5)
        except Exception:
            pass
    # Write new blocks
    lines = content.split('\n') if content.strip() else []
    if not lines:
        return
    children = [
        {'object': 'block', 'type': 'paragraph',
         'paragraph': {'rich_text': [{'type': 'text', 'text': {'content': ln}}] if ln else []}}
        for ln in lines
    ]
    r = requests.patch(
        f'{NOTION_BASE}/blocks/{page_id}/children',
        headers=hdrs, json={'children': children}, timeout=15
    )
    r.raise_for_status()


# ── Docage helpers ────────────────────────────────────────────────────────────

def _docage_auth():
    return (DOCAGE_EMAIL, DOCAGE_API_KEY)


def _docage_headers():
    return {'Accept': 'application/json', 'Content-Type': 'application/json'}


def relancer_docage(email):
    email_lower = email.strip().lower()
    auth = _docage_auth()
    hdrs = _docage_headers()
    try:
        boxes = requests.get(f'{DOCAGE_BASE}/Boxes', auth=auth, headers=hdrs, timeout=10).json()
        if not isinstance(boxes, list):
            return False
    except Exception:
        return False
    for box in boxes:
        box_id = box.get('Id', '')
        if not box_id:
            continue
        try:
            entries = requests.get(
                f'{DOCAGE_BASE}/Boxes/BoxTransactionBatchEntries/{box_id}',
                auth=auth, headers=hdrs, timeout=10
            ).json()
            if not isinstance(entries, list):
                continue
        except Exception:
            continue
        for entry in entries:
            contact_id = entry.get('ContactId', '')
            if not contact_id:
                continue
            try:
                contact = requests.get(
                    f'{DOCAGE_BASE}/Contacts/ById/{contact_id}',
                    auth=auth, headers=hdrs, timeout=10
                ).json()
                if (contact.get('Email') or '').strip().lower() != email_lower:
                    continue
            except Exception:
                continue
            transaction_id = entry.get('TransactionId', '')
            if not transaction_id:
                continue
            try:
                requests.post(
                    f'{DOCAGE_BASE}/Transactions/{transaction_id}/Send',
                    auth=auth, headers=hdrs, timeout=10
                ).raise_for_status()
                return True
            except Exception:
                return False
    return False


# ── Routes : auth ─────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '')
        password = request.form.get('password', '')
        if username == DASHBOARD_USER and password == DASHBOARD_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        error = 'Identifiants incorrects.'
    return render_template('login.html', error=error)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        user_email = os.getenv('DASHBOARD_USER_EMAIL', '').lower()
        if email == user_email:
            token = get_serializer().dumps(email, salt='password-reset')
            reset_url = url_for('reset_password', token=token, _external=True)
            _send_reset_email(user_email, reset_url)
        return render_template('forgot_password.html', sent=True)
    return render_template('forgot_password.html', sent=False)


def _send_reset_email(to_email, reset_url):
    smtp_user     = os.getenv('IMAP_EMAIL', '')
    smtp_password = os.getenv('IMAP_PASSWORD', '')
    body = (
        "Bonjour,\n\n"
        "Vous avez demandé la réinitialisation de votre mot de passe pour Mon Adjoint IA.\n\n"
        f"Cliquez sur ce lien pour définir un nouveau mot de passe (valable 1 heure) :\n{reset_url}\n\n"
        "Si vous n'avez pas fait cette demande, ignorez cet email.\n\n"
        "— Sophie Boutemy"
    )
    msg = MIMEText(body, 'plain', 'utf-8')
    msg['Subject'] = 'Réinitialisation de mot de passe — Mon Adjoint IA'
    msg['From']    = f'Mon Adjoint IA <{smtp_user}>'
    msg['To']      = to_email
    with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
        server.login(smtp_user, smtp_password)
        server.send_message(msg)


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        get_serializer().loads(token, salt='password-reset', max_age=3600)
    except SignatureExpired:
        return render_template('reset_password.html',
                               error='Ce lien a expiré. Faites une nouvelle demande.', token=None)
    except BadSignature:
        return render_template('reset_password.html', error='Lien invalide.', token=None)
    if request.method == 'POST':
        password = request.form.get('password', '')
        confirm  = request.form.get('confirm', '')
        if len(password) < 8:
            return render_template('reset_password.html', token=token,
                                   error='Le mot de passe doit contenir au moins 8 caractères.')
        if password != confirm:
            return render_template('reset_password.html', token=token,
                                   error='Les mots de passe ne correspondent pas.')
        global DASHBOARD_PASSWORD
        DASHBOARD_PASSWORD = password
        set_key(DOTENV_PATH, 'DASHBOARD_PASSWORD', password)
        return render_template('reset_password.html', success=True, token=None)
    return render_template('reset_password.html', token=token)


# ── Routes : dashboard ────────────────────────────────────────────────────────

_AGENT_CARDS = [
    {
        'title':     'Adjoint Client',
        'agent_key': 'Suivi Élèves',
        'btn_label': 'Ouvrir la base élèves',
        'btn_endpoint': 'eleves',
    },
    {
        'title':     'Adjoint Factures',
        'agent_key': 'Factures',
        'btn_label': 'Ouvrir les factures',
        'btn_endpoint': 'factures',
    },
    {
        'title':     "Adjoint Attente",
        'agent_key': "Liste d'attente",
        'btn_label': "Ouvrir la liste d'attente",
        'btn_endpoint': 'liste_attente',
    },
    {
        'title':       'Adjoint Social',
        'agent_key':   '',
        'btn_label':   'Ouvrir Brief to Post',
        'btn_endpoint': 'brief_to_post',
        'description': "Génération de contenus réseaux sociaux et newsletter via IA (Instagram, Facebook, LinkedIn, Newsletter) à partir d'un brief texte — avec historique et édition inline.",
    },
    {
        'title':       'Adjoint Planning',
        'agent_key':   'Cours récurrents',
        'btn_label':   'Voir les cours récurrents',
        'btn_endpoint': 'recurrence',
    },
    {
        'title':       'Adjoint Prospection',
        'agent_key':   '',
        'btn_label':   'Ouvrir la prospection',
        'btn_endpoint': 'prospection',
        'description': "Recherche d'entreprises par secteur et d?partement, enrichissement automatique (site web, email), gestion des contacts avec statuts et notes, r?daction d'emails par IA et suivi des envois.",
    },
]


@app.route('/')
@login_required
def index():
    with open(AUTOMATIONS_CONFIG, encoding='utf-8') as f:
        automations = json.load(f)

    by_agent = {}
    for a in automations:
        if a.get('trigger') == 'systemd':
            status = _systemd_status(a['service'])
        else:
            status = _last_execution(a['log'], a.get('stale_after_hours', 24))
        agent = a.get('agent', '')
        by_agent.setdefault(agent, []).append({'name': a['name'], 'status': status})

    cards = []
    for cfg in _AGENT_CARDS:
        cards.append({
            'title':       cfg['title'],
            'scenarios':   by_agent.get(cfg['agent_key'], []) if cfg['agent_key'] else [],
            'description': cfg.get('description', ''),
            'btn_label':   cfg['btn_label'],
            'btn_url':     url_for(cfg['btn_endpoint']) if cfg['btn_endpoint'] else None,
        })

    return render_template('index.html', cards=cards)


# ── Routes : élèves — liste ───────────────────────────────────────────────────

@app.route('/eleves')
@login_required
def eleves():
    try:
        data = fetch_eleves()
        opts = fetch_db_select_options()
        error = None
    except Exception as e:
        data, opts, error = [], {'boite_mail': [], 'statut_contrat': [], 'type_eleve': []}, str(e)
    return render_template('eleves.html', eleves=data, opts=opts, error=error)


@app.route('/eleves/export.xlsx')
@login_required
def eleves_export():
    statut_filter = request.args.get('statut', '')
    boite_filter  = request.args.get('boite', '')
    type_filter   = request.args.get('type', '')

    data = fetch_eleves()
    if statut_filter:
        data = [e for e in data if e['statut_contrat'] == statut_filter]
    if boite_filter:
        data = [e for e in data if e['boite_mail'] == boite_filter]
    if type_filter:
        data = [e for e in data if type_filter in e['type_eleve']]

    columns = [
        ('Email',                   'email'),
        ('Nom complet',             'nom'),
        ("Type d'élève",            'type_eleve'),
        ('Rattrapage',              'rattrapage'),
        ('Infos',                   'infos'),
        ('Boîte mail',              'boite_mail'),
        ('Date du mail',            'date_mail'),
        ('Infos Calendly',          'infos_calendly'),
        ('Résumé du mail',          'resume_mail'),
        ('Date newsletter envoyée', 'date_newsletter'),
        ('Date contrat envoyé',     'date_contrat'),
        ('Statut contrat envoyé',   'statut_contrat'),
        ('Date de relance',         'date_relance'),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Élèves'
    hfont  = Font(bold=True, color='FFFFFF')
    hfill  = PatternFill('solid', fgColor='354626')
    halign = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx, (col_name, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign
    ws.row_dimensions[1].height = 28
    for row_idx, eleve in enumerate(data, 2):
        for col_idx, (_, key) in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=eleve.get(key, ''))
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 22
    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='eleves.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Routes : élèves — édition inline ─────────────────────────────────────────

@app.route('/eleves/<page_id>/property', methods=['PATCH'])
@login_required
def eleve_update_property(page_id):
    data  = request.get_json(silent=True) or {}
    field = data.get('field', '')
    value = data.get('value', '')
    if field not in FIELD_MAP:
        return jsonify({'ok': False, 'error': f'Champ non éditable : {field}'}), 400
    notion_name, field_type = FIELD_MAP[field]
    prop = _build_notion_prop(field_type, value)
    if prop is None:
        return jsonify({'ok': False, 'error': 'Type non supporté'}), 400
    try:
        r = requests.patch(
            f'{NOTION_BASE}/pages/{page_id}',
            headers=_notion_headers(),
            json={'properties': {notion_name: prop}},
            timeout=10
        )
        r.raise_for_status()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ── Routes : élèves — fiche ───────────────────────────────────────────────────

@app.route('/eleves/<page_id>', methods=['GET', 'DELETE'])
@login_required
def eleve_detail(page_id):
    if request.method == 'DELETE':
        try:
            r = requests.patch(
                f'{NOTION_BASE}/pages/{page_id}',
                headers=_notion_headers(),
                json={'archived': True},
                timeout=10
            )
            r.raise_for_status()
            return jsonify({'ok': True})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)}), 500

    try:
        eleve = fetch_eleve(page_id)
        error = None
    except Exception as e:
        eleve, error = None, str(e)
    return render_template('eleve.html', eleve=eleve, error=error)


@app.route('/eleves/<page_id>/notes', methods=['POST'])
@login_required
def eleve_save_notes(page_id):
    data    = request.get_json(silent=True) or {}
    content = data.get('content', '')
    try:
        save_page_notes(page_id, content)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/eleves/<page_id>/relancer', methods=['POST'])
@login_required
def eleve_relancer(page_id):
    try:
        eleve = fetch_eleve(page_id)
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Notion inaccessible : {e}'}), 500

    if eleve['statut_contrat'] != 'En attente':
        return jsonify({
            'ok': False,
            'error': f'Statut actuel : "{eleve["statut_contrat"]}". Relance possible uniquement pour "En attente".',
        }), 400

    docage_ok = relancer_docage(eleve['email'])
    try:
        update_notion_statut(page_id, 'Relancé')
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Erreur Notion : {e}'}), 500

    msg = ('Relance envoyée via Docage et statut mis à jour dans Notion.' if docage_ok
           else 'Statut mis à jour dans Notion (transaction Docage introuvable).')
    return jsonify({'ok': True, 'message': msg})


# ── Routes : factures ────────────────────────────────────────────────────────

def _page_to_facture(page):
    p = page.get('properties', {})
    # Locate title property dynamically
    nom = ''
    for prop in p.values():
        if prop.get('type') == 'title':
            nom = ''.join(r['plain_text'] for r in prop.get('title', []))
            break
    # Checkbox: try several possible names
    def _checkbox(keys):
        for k in keys:
            if k in p and p[k].get('type') == 'checkbox':
                return p[k].get('checkbox', False)
        return False
    return {
        'id':          page['id'],
        'notion_url':  page.get('url', ''),
        'nom':         nom,
        'date':        _notion_prop(p.get('Date de réception', p.get('Date', {}))),
        'expediteur':  _notion_prop(p.get('Expéditeur', p.get('Expediteur', p.get('De', {})))),
        'comptable':   _checkbox(['Envoyée à la comptable', 'Comptable', 'Envoyée']),
        'lien_drive':  _notion_prop(p.get('Lien Drive', p.get('Drive', p.get('URL', p.get('Fichier', {}))))),
    }


def fetch_factures():
    url = f'{NOTION_BASE}/databases/{NOTION_FACTURES_DB}/query'
    headers = _notion_headers()
    results, cursor = [], None
    while True:
        body = {'page_size': 100, 'sorts': [{'property': 'Date de réception', 'direction': 'descending'}]}
        if cursor:
            body['start_cursor'] = cursor
        r = requests.post(url, headers=headers, json=body, timeout=15)
        if not r.ok:
            body_retry = {'page_size': 100}
            if cursor:
                body_retry['start_cursor'] = cursor
            r = requests.post(url, headers=headers, json=body_retry, timeout=15)
        r.raise_for_status()
        data = r.json()
        results.extend(data.get('results', []))
        if not data.get('has_more'):
            break
        cursor = data.get('next_cursor')
    return [_page_to_facture(p) for p in results]


def fetch_facture(page_id):
    r = requests.get(f'{NOTION_BASE}/pages/{page_id}', headers=_notion_headers(), timeout=10)
    r.raise_for_status()
    return _page_to_facture(r.json())


def update_facture_comptable(page_id, checked):
    # Try to find the correct property name
    schema = requests.get(f'{NOTION_BASE}/databases/{NOTION_FACTURES_DB}', headers=_notion_headers(), timeout=10)
    schema.raise_for_status()
    props = schema.json().get('properties', {})
    checkbox_key = next(
        (k for k, v in props.items() if v.get('type') == 'checkbox'),
        'Envoyée à la comptable'
    )
    body = {'properties': {checkbox_key: {'checkbox': bool(checked)}}}
    r = requests.patch(f'{NOTION_BASE}/pages/{page_id}', headers=_notion_headers(), json=body, timeout=10)
    r.raise_for_status()


@app.route('/factures')
@login_required
def factures():
    try:
        data = fetch_factures()
        error = None
    except Exception as e:
        data, error = [], str(e)
    return render_template('factures.html', factures=data, error=error,
                           drive_url=DRIVE_FACTURES_URL)


@app.route('/factures/export.xlsx')
@login_required
def factures_export():
    expediteur_f = request.args.get('expediteur', '')
    comptable_f  = request.args.get('comptable', '')
    data = fetch_factures()
    if expediteur_f:
        data = [f for f in data if f['expediteur'] == expediteur_f]
    if comptable_f == 'oui':
        data = [f for f in data if f['comptable']]
    elif comptable_f == 'non':
        data = [f for f in data if not f['comptable']]

    columns = [
        ('Nom de la facture',           'nom'),
        ('Date de réception',           'date'),
        ('Expéditeur',                  'expediteur'),
        ('Envoyée à la comptable',      'comptable'),
        ('Lien Drive',                  'lien_drive'),
    ]
    wb = Workbook(); ws = wb.active; ws.title = 'Factures'
    hfont  = Font(bold=True, color='FFFFFF')
    hfill  = PatternFill('solid', fgColor='354626')
    halign = Alignment(horizontal='center', vertical='center')
    for ci, (col_name, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign
    for ri, facture in enumerate(data, 2):
        for ci, (_, key) in enumerate(columns, 1):
            val = facture.get(key, '')
            if isinstance(val, bool):
                val = 'Oui' if val else 'Non'
            ws.cell(row=ri, column=ci, value=val)
    for ci in range(1, len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = 28
    ws.freeze_panes = 'A2'

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='factures.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/factures/refresh', methods=['POST'])
@login_required
def factures_refresh():
    try:
        result = subprocess.run(
            ['python3', FACTURES_SCRIPT],
            capture_output=True, text=True, timeout=90,
            cwd='/home/ubuntu/automations'
        )
        if result.returncode == 0:
            return jsonify({'ok': True, 'message': 'Script exécuté avec succès.'})
        return jsonify({'ok': False, 'error': result.stderr[-500:] or 'Erreur inconnue.'}), 500
    except subprocess.TimeoutExpired:
        return jsonify({'ok': False, 'error': 'Timeout — le script a pris plus de 90 secondes.'}), 500
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/factures/<page_id>/comptable', methods=['PATCH'])
@login_required
def facture_update_comptable(page_id):
    data = request.get_json(silent=True) or {}
    checked = bool(data.get('checked', False))
    try:
        update_facture_comptable(page_id, checked)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


@app.route('/factures/<page_id>')
@login_required
def facture_detail(page_id):
    try:
        facture = fetch_facture(page_id)
        error = None
    except Exception as e:
        facture, error = None, str(e)
    return render_template('facture.html', facture=facture, error=error)


# ── Routes : liste d'attente ──────────────────────────────────────────────────

_MONTH_FR = {
    'Jan': 'jan', 'Feb': 'fév', 'Mar': 'mars', 'Apr': 'avr',
    'May': 'mai', 'Jun': 'juin', 'Jul': 'juil', 'Aug': 'août',
    'Sep': 'sep', 'Oct': 'oct', 'Nov': 'nov', 'Dec': 'déc',
}


def _format_log_ts(ts_str):
    if not ts_str:
        return ''
    try:
        dt = datetime.strptime(ts_str, '%d/%b/%Y %H:%M:%S')
        return f"{dt.day} {_MONTH_FR.get(dt.strftime('%b'), dt.strftime('%b'))} {dt.year} · {dt.strftime('%H:%M')}"
    except Exception:
        return ts_str


def _ts_to_iso(ts_str):
    if not ts_str:
        return ''
    try:
        return datetime.strptime(ts_str, '%d/%b/%Y %H:%M:%S').strftime('%Y-%m-%d')
    except Exception:
        return ''


def _extract_log_person(text):
    """Extract (nom, email) from a log line containing 'NAME <email>'."""
    import re as _re
    m = _re.search(r'<([^>@\s]+@[^>\s]+)>', text)
    if not m:
        return '', ''
    email = m.group(1).strip()
    before = text[:m.start()].strip()
    for sep in (' à ', ': ', ' de '):
        idx = before.rfind(sep)
        if idx >= 0:
            before = before[idx + len(sep):]
            break
    return before.strip(), email


def _parse_waitlist_log(path, max_groups=40):
    """Parse waitlist log into semantic groups: inscription / annulation."""
    try:
        with open(path, encoding='utf-8') as f:
            raw = f.readlines()
    except FileNotFoundError:
        return []
    except Exception:
        return []

    import re as _re
    TS_RE = _re.compile(r'\[(\d{2}/\w{3}/\d{4} \d{2}:\d{2}:\d{2})\]')
    last_ts = None
    tagged = []

    for line in raw:
        line = line.strip()
        if not line:
            continue

        m = TS_RE.search(line)
        if m:
            last_ts = m.group(1)

        if (line.startswith('127.0.0.1') or line.startswith(' * ')
                or 'WARNING' in line or '\x1b[' in line
                or 'Serving Flask' in line or 'Debug mode' in line
                or 'Press CTRL+C' in line or 'Running on' in line):
            continue

        if '[tally] Ajout' in line:
            cat = 'ajout'
        elif '[tally] Erreur' in line:
            cat = 'erreur'
        elif '[calendly] Annulation' in line:
            cat = 'annulation'
        elif '[notifier] Notification envoyée' in line:
            cat = 'notification'
        elif '[notifier] Confirmation envoyée' in line:
            cat = 'confirmation'
        elif '[notifier] Admin notifie' in line:
            cat = 'admin'
        else:
            continue

        tagged.append({'ts': last_ts, 'cat': cat, 'text': line})

    groups = []
    i = 0
    while i < len(tagged):
        ev = tagged[i]
        cat = ev['cat']

        if cat == 'ajout':
            nom, email = _extract_log_person(ev['text'])
            group = {
                'type': 'inscription',
                'ts': _format_log_ts(ev['ts']),
                'ts_iso': _ts_to_iso(ev['ts']),
                'nom': nom,
                'email': email,
                'has_error': False,
            }
            i += 1
            while i < len(tagged) and tagged[i]['cat'] in ('erreur', 'confirmation', 'admin'):
                if tagged[i]['cat'] == 'erreur':
                    group['has_error'] = True
                i += 1
            groups.append(group)

        elif cat == 'annulation':
            m_count = _re.search(r'(\d+) personne', ev['text'])
            count = int(m_count.group(1)) if m_count else 0
            group = {
                'type': 'annulation',
                'ts': _format_log_ts(ev['ts']),
                'ts_iso': _ts_to_iso(ev['ts']),
                'count': count,
                'notifies': [],
            }
            i += 1
            while i < len(tagged) and tagged[i]['cat'] == 'notification':
                nom, email = _extract_log_person(tagged[i]['text'])
                group['notifies'].append({'nom': nom, 'email': email})
                i += 1
            groups.append(group)

        else:
            i += 1

    return list(reversed(groups[-max_groups:]))


@app.route('/liste-attente')
@login_required
def liste_attente():
    try:
        with open(WAITLIST_JSON, encoding='utf-8') as f:
            waitlist = json.load(f)
        if not isinstance(waitlist, list):
            waitlist = []
    except FileNotFoundError:
        waitlist = []
    except Exception:
        waitlist = []

    log_events = _parse_waitlist_log(WAITLIST_LOG)
    service_status = _systemd_status(WAITLIST_SERVICE)

    return render_template('liste_attente.html',
                           waitlist=waitlist,
                           log_events=log_events,
                           service_status=service_status)


# ── Brief to Post : SQLite helper ────────────────────────────────────────────

def _btp_db():
    conn = sqlite3.connect(BTP_DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('''CREATE TABLE IF NOT EXISTS posts (
        id         TEXT PRIMARY KEY,
        titre      TEXT DEFAULT '',
        brief      TEXT NOT NULL,
        ton        TEXT NOT NULL,
        reseau     TEXT NOT NULL,
        texte      TEXT DEFAULT '',
        statut     TEXT DEFAULT 'Généré',
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.commit()
    return conn


# ── Routes : Brief to Post ────────────────────────────────────────────────────

@app.route('/brief-to-post')
@login_required
def brief_to_post():
    reseau    = request.args.get('reseau', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()

    query  = 'SELECT * FROM posts WHERE 1=1'
    params = []
    if reseau:
        query += ' AND reseau = ?'
        params.append(reseau)
    if date_from:
        query += ' AND date(created_at) >= date(?)'
        params.append(date_from)
    if date_to:
        query += ' AND date(created_at) <= date(?)'
        params.append(date_to)
    query += ' ORDER BY created_at DESC'

    conn = _btp_db()
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return render_template('brief_to_post.html', posts=[dict(r) for r in rows],
                           f_reseau=reseau, f_date_from=date_from, f_date_to=date_to)


@app.route('/brief-to-post/generate', methods=['POST'])
@login_required
def brief_to_post_generate():
    data   = request.get_json(silent=True) or {}
    titre  = data.get('titre', '').strip()
    brief  = data.get('brief', '').strip()
    ton    = data.get('ton', '').strip()
    reseau = data.get('reseau', '').strip()

    if not brief or not ton or not reseau:
        return jsonify({'ok': False, 'error': 'Champs obligatoires manquants (brief, ton, réseau)'}), 400

    brief_with_titre = f"Titre : {titre}\n\n{brief}" if titre else brief

    try:
        client  = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        system  = BTP_SYSTEM_PROMPT.format(reseau=reseau, brief=brief_with_titre, ton=ton)
        message = client.messages.create(
            model='claude-sonnet-4-6',
            max_tokens=2048,
            system=system,
            messages=[{'role': 'user', 'content': 'Génère le contenu.'}],
        )
        texte  = message.content[0].text
        statut = 'Généré'
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

    post_id = str(uuid.uuid4())
    conn = _btp_db()
    conn.execute(
        'INSERT INTO posts (id, titre, brief, ton, reseau, texte, statut) VALUES (?, ?, ?, ?, ?, ?, ?)',
        (post_id, titre, brief, ton, reseau, texte, statut),
    )
    conn.commit()
    conn.close()

    return jsonify({
        'ok': True, 'id': post_id, 'texte': texte,
        'titre': titre, 'brief': brief, 'ton': ton, 'reseau': reseau,
    })


@app.route('/brief-to-post/<post_id>', methods=['PATCH', 'DELETE'])
@login_required
def brief_to_post_item(post_id):
    if request.method == 'DELETE':
        conn = _btp_db()
        conn.execute('DELETE FROM posts WHERE id = ?', (post_id,))
        conn.commit()
        conn.close()
        return jsonify({'ok': True})

    data  = request.get_json(silent=True) or {}
    texte = data.get('texte', '')
    conn  = _btp_db()
    conn.execute('UPDATE posts SET texte = ? WHERE id = ?', (texte, post_id))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


# ── Routes : cours récurrents ──────────────────────────────────────────

def _recurrence_pending_count():
    try:
        conn = sqlite3.connect(RECURRENCE_DB_PATH)
        n = conn.execute("SELECT COUNT(*) FROM pending").fetchone()[0]
        conn.close()
        return n
    except Exception:
        return 0


def _recurrence_confirmed_count():
    count = 0
    for log_path, pattern in [
        (LOG_RECURRENCE_WEBHOOK, '✓ Réservé'),
        (LOG_RECURRENCE_RETRY,   'Réservé avec succès'),
    ]:
        try:
            with open(log_path, encoding='utf-8') as f:
                lines = {l.strip() for l in f if pattern in l and l.strip()}
            count += len(lines)
        except FileNotFoundError:
            pass
    return count


@app.route('/recurrence')
@login_required
def recurrence():
    with open(AUTOMATIONS_CONFIG, encoding='utf-8') as f:
        automations = json.load(f)

    services = []
    for a in automations:
        if a.get('agent') != 'Cours récurrents':
            continue
        if a.get('trigger') == 'systemd':
            status = _systemd_status(a['service'])
        else:
            status = _last_execution(a['log'], a.get('stale_after_hours', 25))
        services.append({'name': a['name'], 'status': status})

    return render_template(
        'recurrence.html',
        tally_url=TALLY_FORM_URL,
        pending_count=_recurrence_pending_count(),
        confirmed_count=_recurrence_confirmed_count(),
        services=services,
    )


@app.route('/recurrence/logs/<which>')
@login_required
def recurrence_logs(which):
    MAX_LINES = 100
    log_configs = {
        'webhook': {
            'title':     'recurrence_calendly.log — Webhook Tally',
            'path':      LOG_RECURRENCE_WEBHOOK,
            'max_lines': MAX_LINES,
        },
        'retry': {
            'title':     'recurrence_retry.log — Retry automatique',
            'path':      LOG_RECURRENCE_RETRY,
            'max_lines': MAX_LINES,
        },
    }
    cfg = log_configs.get(which)
    if not cfg:
        return "Log inconnu", 404

    try:
        with open(cfg['path'], encoding='utf-8') as f:
            all_lines = f.readlines()
        deduped = []
        for line in all_lines:
            if not deduped or line != deduped[-1]:
                deduped.append(line)
        lines = [l.rstrip() for l in deduped[-MAX_LINES:]]
    except FileNotFoundError:
        lines = []

    logs = [{'title': cfg['title'], 'lines': lines, 'max_lines': MAX_LINES}]
    return render_template('recurrence_logs.html', logs=logs)


# ── Routes : RGPD — droit à l'effacement ─────────────────────────────────────

def _rgpd_purge_email(email: str) -> dict:
    results = {'db_deleted': 0, 'log_lines': {}, 'errors': []}
    email_lc = email.lower()

    try:
        conn = sqlite3.connect(RECURRENCE_DB_PATH)
        cur  = conn.execute("DELETE FROM pending WHERE lower(email) = ?", (email_lc,))
        results['db_deleted'] = cur.rowcount
        conn.commit()
        conn.close()
    except Exception as e:
        results['errors'].append(f"pending.db : {e}")

    for log_path in RGPD_LOG_FILES:
        fname = os.path.basename(log_path)
        try:
            with open(log_path, encoding='utf-8') as f:
                lines = f.readlines()
            kept    = [l for l in lines if email_lc not in l.lower()]
            removed = len(lines) - len(kept)
            with open(log_path, 'w', encoding='utf-8') as f:
                f.writelines(kept)
            results['log_lines'][fname] = removed
        except FileNotFoundError:
            results['log_lines'][fname] = 0
        except Exception as e:
            results['errors'].append(f"{fname} : {e}")

    return results


def _rgpd_audit_log(operator_ip: str, email: str, results: dict):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    lines = [
        f"{now} INFO [RGPD] Effacement demandé depuis {operator_ip} pour : {email}\n",
        f"{now} INFO [RGPD]   pending.db : {results['db_deleted']} entrée(s) supprimée(s)\n",
    ]
    for fname, n in results.get('log_lines', {}).items():
        lines.append(f"{now} INFO [RGPD]   {fname} : {n} ligne(s) supprimée(s)\n")
    for err in results.get('errors', []):
        lines.append(f"{now} ERROR [RGPD]   Erreur : {err}\n")
    try:
        with open(RGPD_PURGE_LOG, 'a', encoding='utf-8') as f:
            f.writelines(lines)
    except Exception:
        pass


@app.route('/rgpd', methods=['GET', 'POST'])
@login_required
def rgpd():
    if request.method == 'POST':
        email   = request.form.get('email', '').strip()
        confirm = request.form.get('confirm', '')
        if not email or confirm != 'EFFACER':
            return render_template(
                'rgpd.html',
                error="Email requis et confirmation « EFFACER » obligatoire.",
            )
        results = _rgpd_purge_email(email)
        _rgpd_audit_log(request.remote_addr or 'inconnu', email, results)
        return render_template('rgpd.html', done=True, email=email, results=results)
    return render_template('rgpd.html')




# ── Routes Adjoint Prospection ────────────────────────────────────────────────

@app.route('/prospection')
@login_required
def prospection():
    stats = get_stats()
    task = get_task_status()
    today_count = len(get_today_contacts())
    quota = _get_quota()
    profile = load_profile()
    daily_limit = int(profile.get('smtp_daily_limit', 50))
    return render_template('prospection.html', stats=stats, task=task,
                           secteurs=SECTEURS_NAF, statut_options=STATUT_OPTIONS,
                           today_count=today_count, tranche_labels=TRANCHE_LABELS,
                           quota=quota, daily_limit=daily_limit)


@app.route('/prospection/run', methods=['POST'])
@login_required
def prospection_run():
    country = request.form.get('country', 'FR')
    max_results = min(int(request.form.get('max_results', 50)), 200)
    enrichissement = '1' in request.form.getlist('enrichissement')
    if country == 'CH':
        keywords_raw = request.form.get('ch_keywords', '').strip()
        keywords = [k.strip() for k in keywords_raw.splitlines() if k.strip()]
        cantons = request.form.getlist('ch_cantons') or None
        if not keywords:
            return redirect(url_for('prospection'))
        start_search_ch(keywords, cantons, max_results, enrichissement)
    elif country == 'BE':
        keywords_raw = request.form.get('be_keywords', '').strip()
        keywords = [k.strip() for k in keywords_raw.splitlines() if k.strip()]
        provinces = request.form.getlist('be_provinces') or None
        if not keywords:
            return redirect(url_for('prospection'))
        start_search_be(keywords, provinces, max_results, enrichissement)
    else:
        naf_codes = request.form.getlist('naf_codes')
        departements_raw = request.form.get('departements', '').strip()
        departements = [d.strip() for d in departements_raw.split(',') if d.strip()] or None
        tranche_effectifs = request.form.getlist('tranche_effectif') or None
        if not naf_codes:
            return redirect(url_for('prospection'))
        start_search(naf_codes, departements, max_results, enrichissement, tranche_effectifs)
    return redirect(url_for('prospection'))


@app.route('/prospection/status')
@login_required
def prospection_status():
    return jsonify(get_task_status())


@app.route('/prospection/contacts')
@login_required
def prospection_contacts():
    statut  = request.args.get('statut', '')
    naf     = request.args.get('naf', '')
    tranche = request.args.get('tranche', '')
    search  = request.args.get('search', '')
    page    = max(1, int(request.args.get('page', 1)))
    per_page = 50
    rows, total = get_contacts(
        statut=statut or None, naf=naf or None,
        tranche=tranche or None, search=search or None,
        limit=per_page, offset=(page - 1) * per_page,
    )
    naf_list     = get_naf_list()
    tranche_list = get_tranche_list()
    pages = max(1, (total + per_page - 1) // per_page)
    today = datetime.now().strftime('%Y-%m-%d')
    quota = _get_quota()
    profile = load_profile()
    daily_limit = int(profile.get('smtp_daily_limit', 50))
    smtp_configured = bool(profile.get('smtp_user') and profile.get('smtp_pass'))
    return render_template(
        'contacts.html',
        contacts=rows, total=total, page=page, pages=pages,
        naf_list=naf_list, statut_options=STATUT_OPTIONS,
        email_types=EMAIL_TYPES,
        tranche_list=tranche_list, tranche_labels=TRANCHE_LABELS,
        filters={'statut': statut, 'naf': naf, 'tranche': tranche, 'search': search},
        today=today, quota=quota, daily_limit=daily_limit, smtp_configured=smtp_configured,
    )


@app.route('/prospection/contacts/export.csv')
@login_required
def prospection_export():
    statut  = request.args.get('statut', '') or None
    naf     = request.args.get('naf', '') or None
    tranche = request.args.get('tranche', '') or None
    search  = request.args.get('search', '') or None
    rows = get_filtered_for_export(statut=statut, naf=naf, tranche=tranche, search=search)
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=EXPORT_COLS)
    w.writeheader()
    w.writerows(rows)
    buf.seek(0)
    return send_file(
        io.BytesIO(buf.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'contacts_{datetime.now().strftime("%Y%m%d")}.csv',
    )


@app.route('/prospection/contacts/export.xlsx')
@login_required
def prospection_export_xlsx():
    from openpyxl import Workbook as _WB
    statut  = request.args.get('statut', '') or None
    naf     = request.args.get('naf', '') or None
    tranche = request.args.get('tranche', '') or None
    search  = request.args.get('search', '') or None
    rows = get_filtered_for_export(statut=statut, naf=naf, tranche=tranche, search=search)
    wb = _WB(); ws = wb.active; ws.title = 'Contacts'
    ws.append(EXPORT_COLS)
    for row in rows:
        ws.append([row.get(col, '') or '' for col in EXPORT_COLS])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'contacts_{datetime.now().strftime("%Y%m%d")}.xlsx',
    )


@app.route('/prospection/contacts/delete', methods=['POST'])
@login_required
def prospection_contacts_delete():
    ids = [i for i in request.form.getlist('ids') if i.isdigit()]
    if ids:
        delete_contacts([int(i) for i in ids])
    return redirect(request.form.get('next') or url_for('prospection_contacts'))


@app.route('/prospection/contacts/delete-empty', methods=['POST'])
@login_required
def prospection_contacts_delete_empty():
    delete_contacts_empty()
    return redirect(request.form.get('next') or url_for('prospection_contacts'))


@app.route('/prospection/contacts/<int:contact_id>', methods=['PATCH'])
@login_required
def prospection_contact_update(contact_id):
    data = request.get_json(force=True) or {}
    update_contact(contact_id, data)
    return jsonify({'ok': True})


@app.route('/prospection/contacts/<int:contact_id>/desinscrit', methods=['POST'])
@login_required
def prospection_desinscrit(contact_id):
    update_contact(contact_id, {'statut': 'désinscrit'})
    return redirect(request.form.get('next') or url_for('prospection_contacts'))
@app.route('/prospection/unsubscribe/<token>')
def prospection_unsubscribe(token):
    contact_id = _verify_unsub_token(token)
    if contact_id is None:
        html = (
            '<html><head><meta charset="UTF-8"><title>Lien invalide</title>'
            '<style>body{font-family:sans-serif;max-width:520px;margin:80px auto;'
            'padding:0 24px;color:#23242C}a{color:#354626}</style></head>'
            '<body><h2>Lien invalide</h2>'
            '<p>Ce lien de désinscription est invalide.</p>'
            '<p><a href="https://mon-adjoint-ia.fr">mon-adjoint-ia.fr</a></p>'
            '</body></html>'
        )
        return html, 400
    update_contact(contact_id, {'statut': 'désinscrit'})
    html = (
        '<html><head><meta charset="UTF-8"><title>Désinscription confirmée</title>'
        '<style>body{font-family:sans-serif;max-width:520px;margin:80px auto;'
        'padding:0 24px;color:#23242C}a{color:#354626}</style></head>'
        '<body><h2>Désinscription confirmée</h2>'
        '<p>Votre adresse a bien été retirée de notre liste de prospection. '
        'Vous ne recevrez plus aucun message de notre part.</p>'
        '<p><a href="https://mon-adjoint-ia.fr">mon-adjoint-ia.fr</a></p>'
        '</body></html>'
    )
    return html, 200



@app.route('/prospection/contacts/<int:contact_id>/draft', methods=['POST'])
@login_required
def prospection_draft(contact_id):
    contact = get_contact_by_id(contact_id)
    if not contact:
        return jsonify({'error': 'Contact introuvable'}), 404
    data = request.get_json(force=True) or {}
    try:
        token = _make_unsub_token(contact_id)
        unsubscribe_url = url_for('prospection_unsubscribe', token=token, _external=True)
        draft = generate_draft(contact, data.get('email_type', 'premier_contact'), unsubscribe_url=unsubscribe_url)
        return jsonify(draft)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/prospection/contacts/<int:contact_id>/send', methods=['POST'])
@login_required
def prospection_send(contact_id):
    contact = get_contact_by_id(contact_id)
    if not contact:
        return jsonify({'error': 'Contact introuvable'}), 404
    if contact.get('statut') in ('désinscrit', 'archivé'):
        return jsonify({'error': 'Ce contact est désinscrit ou archivé'}), 403
    if not contact.get('email'):
        return jsonify({'error': "Pas d'adresse email pour ce contact"}), 400
    profile = load_profile()
    daily_limit = int(profile.get('smtp_daily_limit', 50))
    if _get_quota() >= daily_limit:
        return jsonify({'error': f'Quota quotidien atteint ({daily_limit} emails/jour)'}), 429
    smtp_user = profile.get('smtp_user', '')
    smtp_pass = profile.get('smtp_pass', '')
    if not smtp_user or not smtp_pass:
        return jsonify({'error': 'SMTP non configuré dans le profil expéditeur'}), 400
    data = request.get_json(force=True) or {}
    subject = data.get('subject', '').strip()
    body    = data.get('body', '').strip()
    if not subject or not body:
        return jsonify({'error': 'Objet et corps requis'}), 400
    smtp_host = profile.get('smtp_host', 'ssl0.ovh.net')
    smtp_port = int(profile.get('smtp_port', 465))
    sender_name = f"{profile.get('prenom', '')} {profile.get('nom', '')}".strip()
    new_statut = 'contacté'
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = f'{sender_name} <{smtp_user}>' if sender_name else smtp_user
        msg['To'] = contact['email']
        _unsub_token = _make_unsub_token(contact_id)
        _unsub_url = url_for('prospection_unsubscribe', token=_unsub_token, _external=True)
        msg['List-Unsubscribe'] = f'<{_unsub_url}>'
        msg['List-Unsubscribe-Post'] = 'List-Unsubscribe=One-Click'
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        msg.attach(MIMEText(_plain_to_html(body), 'html', 'utf-8'))
        with smtplib.SMTP_SSL(smtp_host, smtp_port) as s:
            s.login(smtp_user, smtp_pass)
            s.send_message(msg)
        new_count = _increment_quota()
        note = f"Email envoyé le {datetime.now().strftime('%d/%m/%Y')} : {subject}"
        existing = contact.get('notes', '')
        new_notes = (existing + '\n' + note).strip() if existing else note
        today_str = datetime.now().strftime('%d/%m/%Y')
        date_fields = {'date_premier_contact': today_str} if not contact.get('date_premier_contact') else {'date_relance': today_str}
        update_contact(contact_id, {'statut': new_statut, 'notes': new_notes, **date_fields})
        return jsonify({'ok': True, 'quota': new_count, 'limit': daily_limit})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/prospection/contacts/<int:contact_id>/mark-sent', methods=['POST'])
@login_required
def prospection_mark_sent(contact_id):
    contact = get_contact_by_id(contact_id)
    if not contact:
        return jsonify({'error': 'Contact introuvable'}), 404
    if contact.get('statut') in ('désinscrit', 'archivé'):
        return jsonify({'error': 'Ce contact est désinscrit ou archivé'}), 403
    data = request.get_json(force=True) or {}
    email_type = data.get('email_type', 'premier_contact')
    new_statut = 'contacté'
    type_label = EMAIL_TYPES.get(email_type, email_type)
    note = f"Marqué envoyé le {datetime.now().strftime('%d/%m/%Y')} ({type_label})"
    existing = contact.get('notes', '')
    new_notes = (existing + '\n' + note).strip() if existing else note
    today_str = datetime.now().strftime('%d/%m/%Y')
    date_fields = {'date_premier_contact': today_str} if not contact.get('date_premier_contact') else {'date_relance': today_str}
    update_contact(contact_id, {'statut': new_statut, 'notes': new_notes, **date_fields})
    return jsonify({'ok': True, 'statut': new_statut})


@app.route('/prospection/profile', methods=['GET', 'POST'])
@login_required
def prospection_profile():
    saved = False
    if request.method == 'POST':
        existing = load_profile()
        profile = {
            'prenom':             request.form.get('prenom', '').strip(),
            'nom':                request.form.get('nom', '').strip(),
            'activite':           request.form.get('activite', '').strip(),
            'proposition_valeur': request.form.get('proposition_valeur', '').strip(),
            'offre':              request.form.get('offre', '').strip(),
            'cible':              request.form.get('cible', '').strip(),
            'signature':          request.form.get('signature', '').strip(),
            'smtp_host':          request.form.get('smtp_host', 'ssl0.ovh.net').strip(),
            'smtp_port':          request.form.get('smtp_port', '465').strip(),
            'smtp_user':          request.form.get('smtp_user', '').strip(),
            'smtp_daily_limit':   request.form.get('smtp_daily_limit', '50').strip(),
        }
        new_pass = request.form.get('smtp_pass', '').strip()
        profile['smtp_pass'] = new_pass if new_pass else existing.get('smtp_pass', '')
        save_profile(profile)
        saved = True
    profile = load_profile()
    quota = _get_quota()
    daily_limit = int(profile.get('smtp_daily_limit', 50))
    return render_template('profile.html', profile=profile, saved=saved,
                           quota=quota, daily_limit=daily_limit)


@app.route('/prospection/today')
@login_required
def prospection_today():
    contacts = get_today_contacts()
    today_str = datetime.now().strftime('%d/%m/%Y')
    return render_template('today.html', contacts=contacts,
                           statut_options=STATUT_OPTIONS,
                           email_types=EMAIL_TYPES,
                           today=today_str)


@app.route('/prospection/priorite')
@login_required
def prospection_priorite():
    contacts = get_priorite_contacts(limit=15)
    quota = _get_quota()
    profile = load_profile()
    daily_limit = int(profile.get('smtp_daily_limit', 50))
    smtp_configured = bool(profile.get('smtp_user') and profile.get('smtp_pass'))
    return render_template('priorite.html', contacts=contacts,
                           statut_options=STATUT_OPTIONS,
                           email_types=EMAIL_TYPES,
                           today=datetime.now().strftime('%Y-%m-%d'),
                           quota=quota, daily_limit=daily_limit,
                           smtp_configured=smtp_configured)


@app.route('/prospection/funnel')
@login_required
def prospection_funnel():
    raw = get_funnel_stats()
    total = sum(raw.values()) if raw else 0
    funnel = []
    for val, label in STATUT_OPTIONS:
        funnel.append({'val': val, 'label': label, 'count': raw.get(val, 0)})
    known = {v for v, _ in STATUT_OPTIONS}
    for val, count in raw.items():
        if val not in known:
            funnel.append({'val': val, 'label': val, 'count': count})
    return render_template('funnel.html', funnel=funnel, total=total)


# ── Footer pages ─────────────────────────────────────────────────────────────

@app.context_processor
def inject_year():
    return {'current_year': datetime.now().year}



@app.route('/contact', methods=['GET', 'POST'])
def contact_page():
    sent = False
    error = None
    if request.method == 'POST':
        nom     = request.form.get('nom', '').strip()
        email   = request.form.get('email', '').strip()
        sujet   = request.form.get('sujet', '').strip()
        message = request.form.get('message', '').strip()
        if nom and email and message:
            try:
                smtp_user = os.getenv('IMAP_EMAIL', '')
                smtp_pass = os.getenv('IMAP_PASSWORD', '')
                body = f"Nom : {nom}\nEmail : {email}\n\n{message}"
                msg = MIMEText(body, 'plain', 'utf-8')
                msg['Subject'] = f"[Contact] {sujet or 'Message depuis le dashboard'}"
                msg['From']     = smtp_user
                msg['To']       = 'contact@sophieboutemy.fr'
                msg['Reply-To'] = email
                with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as s:
                    s.login(smtp_user, smtp_pass)
                    s.send_message(msg)
                sent = True
            except Exception as e:
                error = str(e)
        else:
            error = 'Veuillez remplir tous les champs obligatoires.'
    return render_template('contact.html', sent=sent, error=error)


@app.route('/politique-confidentialite')
def politique_confidentialite():
    return render_template('rgpd_info.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5005, debug=False)

